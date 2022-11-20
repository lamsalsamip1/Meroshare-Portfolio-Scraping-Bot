from openpyxl import load_workbook
from datetime import datetime
from read_category import get_categories
from openpyxl.styles import Font


class share_obj():

    def __init__(self, id, name, category, price, balance, ltp, value_ltp, dividend):
        self.id = id
        self.name = name
        self.price = price
        self.balance = balance
        self.dividend = dividend
        self.category = category
        self.ltp = ltp
        self.value_ltp = value_ltp


class old_share():
    def __init__(self, name, price, dividend):
        self.name = name
        self.price = price
        self.dividend = dividend


def get_price(sheet, name, total):

    for r in range(3, total+3):
        s_name = sheet.cell(row=r, column=2).value
        if (s_name == name):
            price = sheet.cell(row=r, column=4).value
            return price
    return 0


def get_dividend(sheet, name, total):
    for r in range(3, total+3):
        s_name = sheet.cell(row=r, column=2).value
        if (s_name == name):
            dividend = sheet.cell(row=r, column=8).value
            return dividend
    return 0


def generate(oldfile, sheet1, sheet2):

    all_items = {}
    current_dateTime = datetime.now()
    wb1 = load_workbook(
        filename=f"./files/Meroshare-{str(current_dateTime)[:10]}.xlsx")
    wb3 = load_workbook(filename=oldfile)
    sheet_write = wb1.get_sheet_by_name(sheet2)  # write sheet
    sheet_data = wb1.get_sheet_by_name(sheet1)  # cdsc sheet

    sheet_report = wb3.get_sheet_by_name(sheet2)  # old file sheet
    categories = get_categories()               # get the categories data
    present_cat = set()
    # create a share object by getting all required data

    total = int(sheet_report['N1'].value)
    for i in range(2, sheet_data.max_row):
        id = sheet_data.cell(row=i, column=1).value
        name = sheet_data.cell(row=i, column=2).value
        balance = sheet_data.cell(row=i, column=3).value
        ltp = sheet_data.cell(row=i, column=6).value
        ltp_value = sheet_data.cell(row=i, column=7).value

        category = categories[name]
        present_cat.add(category)
        price = get_price(sheet_report, name, total)
        dividend = get_dividend(sheet_report, name, total)
        all_items[name] = share_obj(id, name, category, price,
                                    balance, ltp, ltp_value, dividend)

    # write to portfolio file
    k = 3
    for i in all_items:
        sheet_write.cell(row=k, column=1).value = all_items[i].id
        sheet_write.cell(row=k, column=2).value = all_items[i].name
        print(f"{all_items[i].id} {all_items[i].name}")
        sheet_write.cell(row=k, column=3).value = all_items[i].category
        sheet_write.cell(row=k, column=4).value = all_items[i].price
        sheet_write.cell(row=k, column=5).value = all_items[i].balance
        sheet_write.cell(row=k, column=6).value = all_items[i].ltp
        sheet_write.cell(row=k, column=7).value = all_items[i].value_ltp
        sheet_write.cell(row=k, column=8).value = all_items[i].dividend
        sheet_write.cell(row=k, column=9).value = f"=G{k}+H{k}"
        sheet_write.cell(row=k, column=10).value = f"=I{k}-D{k}"
        k = k+1

    # calculate number of items
    total = k-3

    sheet_write['N1'] = total
    if (sheet2 == 'Portfolio_Hom'):
        sheet_write.cell(row=k+3, column=3).value = "All 4 persons"
        sheet_write.cell(row=k+3, column=3).font = Font(bold=True)

    # calculate total of all columns
    for j in range(4, 13):
        sheet_write.cell(
            row=k+1, column=j).value = f"=SUM({chr(64+j)}3:{chr(64+j)}{total+2})"
        sheet_write.cell(
            row=k+1, column=j).font = Font(bold=True)

    # calculate total depending on categories
    new_ind = k+6
    for item in present_cat:
        sheet_write.cell(
            row=new_ind, column=3).value = item
        for j in range(4, 13):
            letter = chr(64+j)
            sheet_write.cell(
                row=new_ind, column=j).value = f"=SUMIF($C$3:$C${total+2},\"{item}\",{letter}$3:{letter}${total+2})"
        new_ind = new_ind+1

    wb1.save(f"./files/Meroshare-{str(current_dateTime)[:10]}.xlsx")


#generate("Hom_CDSC", "Portfolio_Hom")
def return_row(total, sheet, i, cat):
    for r in range(total+5, total+15):
        cat_value = sheet.cell(row=r, column=3).value
        if (cat_value == cat):
            return r
    return 0


def return_formula(samip, gita, sharad, letter):
    org = f"Portfolio_Gita!{letter}{gita}+Portfolio_Sharad!{letter}{sharad}+Portfolio_Samip!{letter}{samip}"
    if samip == 0:
        org = org.replace(f"Portfolio_Samip!{letter}{samip}", '0')
    if gita == 0:
        org = org.replace(f"Portfolio_Gita!{letter}{gita}", '0')
    if sharad == 0:
        org = org.replace(f"Portfolio_Sharad!{letter}{sharad}", '0')
    return org


def generate_overall():

    current_dateTime = datetime.now()
    workbook = load_workbook(
        f"./files/Meroshare-{str(current_dateTime)[:10]}.xlsx")

    hom = workbook.get_sheet_by_name("Portfolio_Hom")
    samip = workbook.get_sheet_by_name("Portfolio_Samip")
    sharad = workbook.get_sheet_by_name("Portfolio_Sharad")
    gita = workbook.get_sheet_by_name("Portfolio_Gita")

    total_hom_row = int(hom['N1'].value)+4
    total_gita_row = int(gita['N1'].value)+4
    total_sharad_row = int(sharad['N1'].value)+4
    total_samip_row = int(samip['N1'].value)+4
    print(total_gita_row, total_hom_row, total_samip_row, total_sharad_row)

    # generate overall total
    for j in range(4, 13):
        letter = chr(64+j)
        print(f"={letter}{total_hom_row}+Portfolio_Gita!{letter}{total_gita_row}+Portfolio_Sharad!{letter}{total_sharad_row}+Portfolio_Samip!{letter}{total_samip_row}")
        hom.cell(row=total_hom_row+2,
                 column=j).value = f"={letter}{total_hom_row}+Portfolio_Gita!{letter}{total_gita_row}+Portfolio_Sharad!{letter}{total_sharad_row}+Portfolio_Samip!{letter}{total_samip_row}"

    # for categories
    hom.cell(row=total_hom_row+17, column=3).value = "All 4 persons"
    hom.cell(row=total_hom_row+17, column=3).font = Font(bold=True)

    for i in range(10):
        sum_gita = []
        sum_samip = []
        sum_sharad = []

        cat = hom.cell(row=total_hom_row+5+i, column=3).value
        hom.cell(row=total_hom_row+19+i, column=3).value = cat
        gita_ind = return_row(total_gita_row, gita, i, cat)
        sharad_ind = return_row(total_sharad_row, sharad, i, cat)
        samip_ind = return_row(total_samip_row, samip, i, cat)

        for j in range(4, 13):
            letter = chr(64+j)
            formula = f"={letter}{total_hom_row+5+i}+{return_formula(samip_ind, gita_ind, sharad_ind, letter)}"
            print(formula)
            hom.cell(row=total_hom_row+19+i, column=j).value = formula

    workbook.save(f"./files/Meroshare-{str(current_dateTime)[:10]}.xlsx")
