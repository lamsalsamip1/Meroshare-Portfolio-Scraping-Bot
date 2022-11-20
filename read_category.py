def get_categories():
    from openpyxl import load_workbook
    workbook = load_workbook(filename="./Categories/category.xlsx")
    ws = workbook.active
    categories = {}

    for i in range(1, ws.max_row+1):
        name = ws.cell(row=i, column=1).value
        category = ws.cell(row=i, column=2).value
        categories[name] = category

    categories['TJVCL'] = "Hydropower"
    categories['SGI'] = "Non-Life Insurance"
    return categories
