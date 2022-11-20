from selenium import webdriver
import os
import sys
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.select import Select
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from datetime import datetime
import shutil
# WEBDRIVER INITIALIZATION


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


def initDriver():
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.headless = True
    global driver
    driver = webdriver.Chrome(resource_path(
        './driver.exe'), options=options)


def login(branch_code, username, password):
    driver.get('https://meroshare.cdsc.com.np/')
    # WAIT
    driver.implicitly_wait(5)
    WebDriverWait(driver, 30).until(
        EC.presence_of_element_located((By.TAG_NAME, "app-login")))

    # SELECTORS
    user_id = driver.find_element(By.XPATH, '//*[@id="username"]')
    pw_id = driver.find_element(By.XPATH, '//*[@id="password"]')
    branch = driver.find_element(By.XPATH, '//*[@id="selectBranch"]')

    # SELECT THE BRANCH
    branch.click()
    branch_id = driver.find_element(By.CLASS_NAME, "select2-search__field")
    branch_id.click()
    branch_id.send_keys(branch_code)
    branch_id.send_keys(Keys.ENTER)

    # SEND USERNAME AND PASSWORD
    user_id.send_keys(username)
    pw_id.send_keys(password)

    # WAIT 2 SECONDS
    time.sleep(2)

    # CLICK LOGIN
    login_btn = driver.find_element(By.CLASS_NAME, 'sign-in')
    login_btn.click()

    # CLICK ON PORTFOLIO
    btn_port = driver.find_element(
        By.XPATH, '//*[@id="sideBar"]/nav/ul/li[5]/a')
    btn_port.click()


def collectData(sheet):

    # COLLECT DATA
    rows = 1 + len(
        driver.find_elements(By.XPATH, "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody[1]/tr"))
    cols = 1 + len(
        driver.find_elements(By.XPATH,
                             "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody[1]/tr[1]/td"))

    for r in range(1, rows):
        for p in range(1, cols):
            location = "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody/tr[" + str(
                r) + "]/td[" + str(
                p) + "]"
            value = driver.find_element(By.XPATH, location).text
            #print(value, end=' ')
            if (p != 2):
                value = float(value.replace(',', ''))
            #print(f"{value} ")
            sheet.cell(row=r + 1, column=p).value = value

    for i in range(2):
        location = f"//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody[2]/tr/td[{(i+1)*2}]"
        value = driver.find_element(By.XPATH, location).text[3:]
        value = float(value.replace(',', ''))
        sheet.cell(row=rows+1, column=3+2*(i+1)).value = value


def scrape(username, password, sheet_name, file, sheet2_name):

    branch_code = 11200

    # DRIVER INITIALIZATION
    initDriver()

    # LOGIN
    login(branch_code, username, password)

    workbook = load_workbook(filename=file)

    sheet = workbook.get_sheet_by_name(sheet_name)

    # count old data
    c = 1
    for item in range(1, sheet.max_row):
        c = c+1

    current_dateTime = datetime.now()
    # DATA COLLECTION
    collectData(sheet)
    sheet2 = workbook.get_sheet_by_name(sheet2_name)
    sheet2['A1'] = str(current_dateTime)[:10]

    # SAVE TO EXCEL SHEET

    workbook.save(filename=file)
    driver.quit()
    return c

    # elif mode == 2:
    #     while True:
    #         time.sleep(10)
    #         driver.refresh()
    #         collectData()
    #         workbook.save(filename=excelFile)
    #         print("DATA UPDATED : "+str(datetime.now()))


# print("*****************************MENU**********************")
# print("1. Manually update data")
# print("2. Turn on Automatic Updating")
# choice = int(input("Enter your input (1/2): "))
