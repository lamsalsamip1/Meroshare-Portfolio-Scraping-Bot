from selenium import webdriver
import os
import time
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime

# WEBDRIVER INITIALIZATION
os.environ['PATH'] += r"C:/Selenium drivers"
driver = webdriver.Chrome()


#WORKBOOK INITIALIZATION
workbook = load_workbook(filename="excelfile.xlsx")
sheet = workbook.active


def login(branch_code, username, password):
    driver.get('https://meroshare.cdsc.com.np/')
    #WAIT
    driver.implicitly_wait(5)
    WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.TAG_NAME, "app-login")))

    #SELECTORS
    user_id = driver.find_element(By.XPATH, '//*[@id="username"]')
    pw_id = driver.find_element(By.XPATH, '//*[@id="password"]')
    branch = driver.find_element(By.XPATH, '//*[@id="selectBranch"]')

    # SELECT THE BRANCH
    branch.click()
    branch_id = driver.find_element(By.CLASS_NAME, "select2-search__field")
    branch_id.click()
    branch_id.send_keys(branch_code)
    branch_id.send_keys(Keys.ENTER)

    # SEND USERNAME AND PASSWORD
    user_id.send_keys(username)
    pw_id.send_keys(password)

    # WAIT 2 SECONDS
    time.sleep(2)

    # CLICK LOGIN
    login_btn = driver.find_element(By.CLASS_NAME, 'sign-in')
    login_btn.click()

    # CLICK ON PORTFOLIO
    btn_port = driver.find_element(By.XPATH, '//*[@id="sideBar"]/nav/ul/li[5]/a')
    btn_port.click()


def collectData():


    # FIND NUMBER OF ROWS AND COLUMNS IN PORTFOLIO TABLE
    rows = 1 + len(
        driver.find_elements(By.XPATH, "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody[1]/tr"))
    cols = 1 + len(
        driver.find_elements(By.XPATH,
                             "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody[1]/tr[1]/td"))

    # LOOP THROUGH THE TABLE, SCRAPE DATA AND WRITE IT TO EXCEL FILE
    for r in range(1, rows):
        for p in range(1, cols):
            location = "//*[@id='main']/div/app-my-portfolio/div/div[2]/div/div/table/tbody/tr[" + str(
                r) + "]/td[" + str(
                p) + "]"
            value = driver.find_element(By.XPATH, location).text
            print(value, end=' ')

            # REMOVE COMMA IN VALUES
            if (p != 2):
                value = float(value.replace(',', ''))

            # WRITE TO FILE
            sheet.cell(row=r + 1, column=p).value = value
        print()


def init(mode):

    # ALL NEEDED VARIABLES
    excelFile = "excelfile.xlsx"
    username = xxxx
    password = "xxxxx"
    branch_code = xxxxx

    #LOGIN
    login(branch_code, username, password)

    # DATA COLLECTION
    collectData()

    #SAVE TO EXCEL SHEET
    workbook.save(filename=excelFile)
    print(mode)
    if mode == 1:
        print("DATA UPDATED....NOW EXITING")
        driver.quit()
        exit()

    #AUTOMATIC RELOAD AND SCRAPE DATA EVERY 10 MINUTES LOGIC
    elif mode == 2:
        while True:
            time.sleep(600)
            driver.refresh();
            collectData()
            workbook.save(filename=excelFile)
            print("DATA UPDATED : "+str(datetime.now()))

#USER MENU TO UPDATE AND CLOSE OR KEEP PROGRAM IN BACKGROUND FOR AUTOMATIC UPDATE

print("*****************************MENU**********************")
print("1. Manually update data")
print("2. Turn on Automatic Updating")
choice = int(input("Enter your input (1/2): "))

init(choice)



