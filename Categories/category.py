from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import os
from selenium.webdriver.common.by import By
import time
import sys
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook


def resource_path(relative_path):
    try:
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(__file__)
    return os.path.join(base_path, relative_path)


def initDriver():
    options = webdriver.ChromeOptions()
    options.add_experimental_option('excludeSwitches', ['enable-logging'])
    # options.headless = True
    global driver
    driver = webdriver.Chrome(resource_path(
        '..\/driver.exe'), options=options)


def find_length(text):
    words = text.split()
    print(words)
    if (int(words[3]) == 0 and int(words[1]) == 0):
        return 0
    len = int(words[3])-int(words[1])+1
    return len


def scrape_data():

    initDriver()
    workbook = load_workbook(filename="category.xlsx")
    ws = workbook.active
    index = 0
    driver.get('https://www.sharesansar.com/company-list')
    categories = ["Commercial Bank", "Corporate Debentures", "Development Bank", "Finance", "Government Bonds", "Hotel & Tourism", "Hydropower", "Investment",
                  "Life Insurance", "Manufacturing and Products", "Microfinance", "Mutual Fund", "Non-Life Insurance", "Trading", "Promoter Share", "Preference Share", "Others"]
    companies = {}

    for item in categories:

        dropdown = driver.find_element(
            By.CLASS_NAME, 'select2-selection__rendered')
        search_btn = driver.find_element(By.ID, 'btn_listed_submit')
        dropdown.click()
        input = driver.find_element(By.CLASS_NAME, 'select2-search__field')
        input.send_keys(item)
        search_btn.click()

        time.sleep(2)
        while True:

            element = driver.find_element(By.XPATH, '//*[@id="myTable_info"]')
            length = find_length(element.text)
            if (length > 0):
                for comp in range(1, length+1):
                    index = index+1
                    location = f"//*[@id='myTable']/tbody/tr[{comp}]/td[2]/a"
                    name = driver.find_element(By.XPATH, location).text
                    companies[name] = item
                    ws.cell(row=index, column=1).value = name
                    ws.cell(row=index, column=2).value = item
                    print(f"{name}->{item}")

                next_btn = driver.find_element(
                    By.ID, 'myTable_next')

                if "disabled" in next_btn.get_attribute('class'):
                    break
                else:
                    next_btn.click()
                    time.sleep(3)
            else:
                break
        # driver.refresh()
        driver.find_element(By.TAG_NAME, 'body').send_keys(
            Keys.CONTROL + Keys.HOME)
        time.sleep(2)
    workbook.save("category.xlsx")
    driver.quit()
    print(companies)


scrape_data()
